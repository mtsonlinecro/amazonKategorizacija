from __future__ import annotations

import csv
import json
import os
import re
import sqlite3
import tempfile
import threading
import unicodedata
import urllib.request
from urllib.parse import parse_qs, urlparse
from datetime import datetime
from difflib import SequenceMatcher
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path
from typing import Any, Dict, List, Tuple

from openpyxl import load_workbook

try:
    import xlrd  # podrška za stare Amazon .xls datoteke
except Exception:
    xlrd = None

try:
    from category_translation import CONCEPT_TRANSLATIONS, CRITICAL_LEAF_CONCEPTS, OPPOSING_CONCEPTS
except Exception:
    CONCEPT_TRANSLATIONS = {}
    CRITICAL_LEAF_CONCEPTS = set()
    OPPOSING_CONCEPTS = {}

HOST = "127.0.0.1"
PORT = 8008
BASE_DIR = Path(__file__).resolve().parent
DATA_DIR = BASE_DIR / "data"
DATA_DIR.mkdir(exist_ok=True)
DB_PATH = DATA_DIR / "learning.db"
TEMP_DIR = Path(tempfile.gettempdir()) / "amazon_category_mapper_backend"
TEMP_DIR.mkdir(parents=True, exist_ok=True)


# Progress se čuva samo dok backend radi. WinForms šalje job_id i periodično čita /api/progress.
PROGRESS_LOCK = threading.Lock()
PROGRESS_STATE: Dict[str, Dict[str, Any]] = {}
CANCELLED_JOBS: set[str] = set()


def cancel_job(job_id: str | None) -> None:
    if not job_id:
        return
    with PROGRESS_LOCK:
        CANCELLED_JOBS.add(job_id)
        state = PROGRESS_STATE.get(job_id, {"job_id": job_id})
        state.update({
            "percent": int(state.get("percent", 0) or 0),
            "message": "Zaustavljanje obrade...",
            "status": "cancelled",
            "updated_at": datetime.utcnow().isoformat(timespec="seconds"),
        })
        PROGRESS_STATE[job_id] = state


def is_cancelled(job_id: str | None) -> bool:
    if not job_id:
        return False
    with PROGRESS_LOCK:
        return job_id in CANCELLED_JOBS


def check_cancelled(job_id: str | None) -> None:
    if is_cancelled(job_id):
        raise RuntimeError("Obrada je zaustavljena na zahtjev korisnika.")


def set_progress(job_id: str | None, percent: int, message: str, status: str = "running") -> None:
    if not job_id:
        return
    percent = max(0, min(100, int(percent)))
    with PROGRESS_LOCK:
        PROGRESS_STATE[job_id] = {
            "job_id": job_id,
            "percent": percent,
            "message": message,
            "status": status,
            "updated_at": datetime.utcnow().isoformat(timespec="seconds"),
        }


def get_progress(job_id: str | None) -> Dict[str, Any]:
    if not job_id:
        return {"job_id": "", "percent": 0, "message": "Nema aktivne obrade.", "status": "missing"}
    with PROGRESS_LOCK:
        return dict(PROGRESS_STATE.get(job_id, {
            "job_id": job_id,
            "percent": 0,
            "message": "Čekam početak obrade...",
            "status": "waiting",
        }))


# Korisnikov ulazni CSV/XLSX format.
# Važno: Node Id in FR/IT/ES = node ID, a FR/IT/ES/NL/PL/IE/SE = naziv/path kategorije. UK je namjerno izbačen.
INPUT_COLUMNS = [
    "product type", "NODE ID", "Amazon category name", "Node Id in FR",
    "Node Id in IT", "Node Id in ES", "category name eng", "PIM category name", "EAN",
    "FR", "IT", "ES", "NL", "PL", "IE", "SE",
]
MARKETPLACES = ["FR", "IT", "ES", "NL", "PL", "IE", "SE"]
DIRECT_MARKETPLACES = {"FR", "IT", "ES"}
AI_FALLBACK_MARKETPLACES = {"NL", "PL", "IE", "SE"}
DIRECT_COLUMNS = {"FR": "Node Id in FR", "IT": "Node Id in IT", "ES": "Node Id in ES"}
REQUIRED_MAPPING_COLUMNS = ["Node root", "Node ID", "Node Path", "Node Id in FR", "Node Id in IT", "Node Id in ES"]
INPUT_REQUIRED_ANY = ["product type", "NODE ID", "Amazon category name", "category name eng", "PIM category name", "EAN"]

# UK je izbačen iz obrade i izlaza. Ako se pojavi u starom ulaznom CSV-u, brišemo ga iz rezultata.
UK_DROP_COLUMNS = {
    "UK", "Node Id in UK", "UK_status", "UK_source", "UK_note", "UK_confidence",
    "UK_category_name", "UK_node_id", "UK_source_url", "input_source",
    "NORD", "Node Id in NORD", "NORD_status", "NORD_source", "NORD_note",
    "NORD_confidence", "NORD_category_name", "NORD_node_id", "NORD_source_url",
}


def drop_unused_marketplace_columns(row: Dict[str, Any]) -> None:
    for col in UK_DROP_COLUMNS:
        row.pop(col, None)


def load_env() -> Dict[str, str]:
    values: Dict[str, str] = {}

    # 1) .env ostaje podržan zbog stare verzije projekta.
    env_path = BASE_DIR / ".env"
    if env_path.exists():
        for line in env_path.read_text(encoding="utf-8", errors="ignore").splitlines():
            line = line.strip()
            if not line or line.startswith("#") or "=" not in line:
                continue
            key, value = line.split("=", 1)
            key = key.strip()
            value = value.strip()
            if value:
                values[key] = value

    # 2) Jednostavnije za izmjenu: backend/ai_settings.py.
    try:
        import ai_settings  # type: ignore
        for key in ["OPENAI_API_KEY", "OPENAI_MODEL", "AI_ENABLED"]:
            value = clean(getattr(ai_settings, key, ""))
            if value and key not in values:
                values[key] = value
    except Exception:
        pass

    # 3) Environment varijable imaju zadnju riječ.
    for k, v in os.environ.items():
        if (k.startswith("OPENAI_") or k == "AI_ENABLED") and v:
            values[k] = v
    return values


def ai_enabled_from_config() -> bool:
    env = load_env()
    return clean(env.get("AI_ENABLED", "")).lower() in {"1", "true", "yes", "da"} and bool(clean(env.get("OPENAI_API_KEY")))


def clean(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    if text.lower() in {"nan", "none", "null"}:
        return ""
    if text.endswith(".0") and text[:-2].isdigit():
        text = text[:-2]
    return text


def normalize_header(value: Any) -> str:
    return clean(value).replace("\ufeff", "").strip()


def normalize_column_key(value: Any) -> str:
    text = normalize_header(value).lower()
    return re.sub(r"[^a-z0-9]+", "", text)


COLUMN_ALIASES = {
    # mapping file
    "noderoot": "Node root",
    "nodeid": "Node ID",
    "nodepath": "Node Path",
    "nodeidinfr": "Node Id in FR",
    "nodeidfr": "Node Id in FR",
    "nodeidinit": "Node Id in IT",
    "nodeidit": "Node Id in IT",
    "nodeidines": "Node Id in ES",
    "nodeides": "Node Id in ES",
    # input file
    "nodeidinput": "NODE ID",
    "node": "NODE ID",
    "producttype": "product type",
    "amazoncategoryname": "Amazon category name",
    "categorynameeng": "category name eng",
    "pimcategoryname": "PIM category name",
    "ean": "EAN",
    # category catalog
    "marketplace": "marketplace",
    "country": "marketplace",
    "store": "marketplace",
    "nodeidcategory": "node_id",
    "browseNodeId": "node_id",
    "browsenodeid": "node_id",
    "browse_node_id": "node_id",
    "node_id": "node_id",
    "categoryname": "category_name",
    "displayname": "category_name",
    "name": "category_name",
    "path": "node_path",
    "nodepathcategory": "node_path",
    "categorypath": "node_path",
    "node_path": "node_path",
}

# Dodaj marketplace stupce kao alias same na sebe
for _mp in MARKETPLACES + ["DE"]:
    COLUMN_ALIASES[normalize_column_key(_mp)] = _mp


def canonical_header(value: Any) -> str:
    raw = normalize_header(value)
    key = normalize_column_key(raw)
    return COLUMN_ALIASES.get(key, raw)


def find_header_row(values: List[Tuple[Any, ...]], required: List[str], min_score: int = 1) -> int:
    if not values:
        return 0
    required_keys = {normalize_column_key(c) for c in required}
    best_idx = 0
    best_score = -1
    for idx, row in enumerate(values[:150]):
        raw_keys = {normalize_column_key(v) for v in row if normalize_header(v)}
        canonical_keys = {normalize_column_key(COLUMN_ALIASES.get(k, k)) for k in raw_keys}
        score = len(required_keys.intersection(canonical_keys))
        if score > best_score:
            best_score = score
            best_idx = idx
        if required_keys and required_keys.issubset(canonical_keys):
            return idx
    return best_idx if best_score >= min_score else 0


def rows_from_matrix(values: List[Tuple[Any, ...]], header_idx: int = 0) -> List[Dict[str, str]]:
    if not values:
        return []
    headers = [canonical_header(v) for v in values[header_idx]]
    rows: List[Dict[str, str]] = []
    for row_values in values[header_idx + 1:]:
        row: Dict[str, str] = {}
        for i, header in enumerate(headers):
            if not header:
                continue
            value = clean(row_values[i] if i < len(row_values) else "")
            row[header] = value
        if any(v for v in row.values()):
            rows.append(row)
    return rows


def read_csv_rows(path: Path, find_header: bool = False, required: List[str] | None = None) -> List[Dict[str, str]]:
    raw = path.read_text(encoding="utf-8-sig", errors="replace")
    sample = raw[:8192]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
    except Exception:
        # Excel u HR često koristi ; kao separator
        dialect = csv.excel
        if sample.count(";") > sample.count(","):
            dialect.delimiter = ";"
    reader = csv.reader(raw.splitlines(), dialect=dialect)
    values = [tuple(r) for r in reader]
    header_idx = find_header_row(values, required or [], min_score=1) if find_header and required else 0
    return rows_from_matrix(values, header_idx)


def read_xlsx_rows(path: Path, sheet_name: str | None = None, find_header: bool = False, required: List[str] | None = None) -> List[Dict[str, str]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    if sheet_name:
        sheet_lookup = {s.lower(): s for s in wb.sheetnames}
        wanted = sheet_lookup.get(sheet_name.lower())
        if not wanted:
            raise ValueError(f"Excel mora imati sheet/tab '{sheet_name}'. Dostupni tabovi: {', '.join(wb.sheetnames)}")
        ws = wb[wanted]
    else:
        ws = wb[wb.sheetnames[0]]
    values = list(ws.iter_rows(values_only=True))
    header_idx = find_header_row(values, required or [], min_score=1) if find_header and required else 0
    return rows_from_matrix(values, header_idx)


def read_xls_rows(path: Path, sheet_name: str | None = None, find_header: bool = False, required: List[str] | None = None) -> List[Dict[str, str]]:
    if xlrd is None:
        raise ValueError("Za .xls treba paket xlrd. Pokreni: .venv\\Scripts\\python -m pip install xlrd==2.0.1")
    book = xlrd.open_workbook(str(path))
    if sheet_name:
        names = book.sheet_names()
        lookup = {n.lower(): n for n in names}
        wanted = lookup.get(sheet_name.lower())
        if not wanted:
            raise ValueError(f"Excel mora imati sheet/tab '{sheet_name}'. Dostupni tabovi: {', '.join(names)}")
        sh = book.sheet_by_name(wanted)
    else:
        sh = book.sheet_by_index(0)
    values: List[Tuple[Any, ...]] = []
    for r in range(sh.nrows):
        values.append(tuple(sh.cell_value(r, c) for c in range(sh.ncols)))
    header_idx = find_header_row(values, required or [], min_score=1) if find_header and required else 0
    return rows_from_matrix(values, header_idx)


def read_generic_rows(path: Path, sheet_name: str | None = None, find_header: bool = False, required: List[str] | None = None) -> List[Dict[str, str]]:
    suffix = path.suffix.lower()
    if suffix in {".xlsx", ".xlsm"}:
        return read_xlsx_rows(path, sheet_name=sheet_name, find_header=find_header, required=required)
    if suffix == ".xls":
        return read_xls_rows(path, sheet_name=sheet_name, find_header=find_header, required=required)
    return read_csv_rows(path, find_header=find_header, required=required)


def normalize_input_row(row: Dict[str, str]) -> Dict[str, str]:
    if not clean(row.get("NODE ID")) and clean(row.get("Node ID")):
        row["NODE ID"] = clean(row.get("Node ID"))
    row.pop("Node ID", None)
    return row


def read_node_id_list_rows(path: Path) -> List[Dict[str, str]]:
    """
    Minimalni ulaz: obični .txt ili CSV bez headera koji sadrži samo njemačke DE node ID-eve.
    Primjer TXT-a:
        3597051031
        316968011

    U fallbacku namjerno čitamo samo brojeve duljine 4-20 znamenki. Koristi se tek kad
    normalni CSV/XLSX parser ne pronađe stupac NODE ID / proizvodne podatke.
    """
    try:
        raw = path.read_text(encoding="utf-8-sig", errors="replace")
    except Exception:
        return []
    ids: List[str] = []
    seen = set()
    for match in re.finditer(r"(?<!\d)(\d{4,20})(?!\d)", raw):
        node_id = match.group(1)
        if node_id not in seen:
            ids.append(node_id)
            seen.add(node_id)
    rows: List[Dict[str, str]] = []
    for node_id in ids:
        row = {col: "" for col in INPUT_COLUMNS}
        row["NODE ID"] = node_id
        row["input_source"] = "NODE_ID_LIST"
        rows.append(row)
    return rows


def read_input_rows(path: Path) -> List[Dict[str, str]]:
    # TXT je sada službeno podržan kao najjednostavniji format: jedan DE node ID po retku.
    if path.suffix.lower() == ".txt":
        rows = read_node_id_list_rows(path)
        if rows:
            return rows

    # Pronađi header red, jer neki exporti imaju uvodne retke.
    try:
        rows = read_generic_rows(path, find_header=True, required=INPUT_REQUIRED_ANY)
    except Exception:
        rows = []

    for row in rows:
        normalize_input_row(row)
        for col in INPUT_COLUMNS:
            row.setdefault(col, "")

    # Makni prazne redove i redove koji očito nisu proizvod.
    rows = [r for r in rows if any(clean(r.get(c)) for c in ["NODE ID", "product type", "Amazon category name", "EAN", "PIM category name"])]
    if rows:
        return rows

    # Ako je korisnik poslao CSV/TXT bez headera i samo node ID-eve, prihvati i to.
    fallback_rows = read_node_id_list_rows(path)
    if fallback_rows:
        return fallback_rows

    raise ValueError("Ulazna datoteka je pročitana, ali nije pronađen nijedan podatkovni red. Dovoljno je da pošalješ CSV/XLSX sa stupcem 'NODE ID' ili TXT/CSV s jednim DE node ID-em po retku.")


def write_csv_rows(rows: List[Dict[str, Any]], path: Path, marketplaces: List[str] | None = None) -> None:
    selected = [m.upper().strip() for m in (marketplaces or MARKETPLACES) if m.upper().strip() in MARKETPLACES]
    selected_set = set(selected)

    def should_skip_column(key: str) -> bool:
        k = clean(key)
        if not k or k in UK_DROP_COLUMNS:
            return True
        low = k.lower()
        if low.endswith("_source") or low.endswith("_source_url") or low == "source" or "source_url" in low:
            return True
        if " in uk" in low or low == "uk" or low.startswith("uk_"):
            return True
        if "nord" in low:
            return True
        for mp in MARKETPLACES:
            if mp not in selected_set:
                if k == mp or k == f"Node Id in {mp}" or k.startswith(f"{mp}_"):
                    return True
        return False

    for row in rows:
        drop_unused_marketplace_columns(row)

    base_headers = [
        "product type",
        "NODE ID",
        "Node Id in DE",
        "Amazon category name",
        "DE",
        "category name eng",
        "PIM category name",
        "EAN",
    ]
    headers: List[str] = []
    for preferred in base_headers:
        if not should_skip_column(preferred) and preferred not in headers:
            headers.append(preferred)

    for mp in selected:
        for col in [f"Node Id in {mp}", mp, f"{mp}_category_name", f"{mp}_status", f"{mp}_note", f"{mp}_confidence"]:
            if not should_skip_column(col) and col not in headers:
                headers.append(col)

    for row in rows:
        for key in row.keys():
            if should_skip_column(key):
                continue
            if key not in headers:
                headers.append(key)

    with path.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=headers, extrasaction="ignore")
        writer.writeheader()
        for row in rows:
            writer.writerow({k: clean(row.get(k, "")) for k in headers})


def read_mapping_rows(path: Path) -> List[Dict[str, str]]:
    rows = read_generic_rows(path, sheet_name="MAPPINGS", find_header=True, required=REQUIRED_MAPPING_COLUMNS)
    if not rows:
        raise ValueError("MAPPINGS tab je prazan ili nije pronađen nijedan podatkovni red.")
    first_keys = set(rows[0].keys())
    missing = [c for c in REQUIRED_MAPPING_COLUMNS if c not in first_keys]
    if missing:
        found = ", ".join(rows[0].keys())
        raise ValueError("U MAPPINGS tabu fale stupci: " + ", ".join(missing) + ". Pronađeni stupci su: " + found)
    return [r for r in rows if clean(r.get("Node ID"))]


def build_mapping_index(mapping_rows: List[Dict[str, str]]) -> Dict[str, Dict[str, str]]:
    index: Dict[str, Dict[str, str]] = {}
    for row in mapping_rows:
        node_id = clean(row.get("Node ID"))
        if node_id:
            index[node_id] = {col: clean(row.get(col)) for col in REQUIRED_MAPPING_COLUMNS}
    return index


def build_category_index(rows: List[Dict[str, str]]) -> Dict[Tuple[str, str], Dict[str, str]]:
    idx: Dict[Tuple[str, str], Dict[str, str]] = {}
    for r in rows:
        mp = clean(r.get("marketplace")).upper()
        node_id = clean(r.get("node_id"))
        if mp and node_id:
            idx[(mp, node_id)] = r
    return idx


def infer_marketplace_from_text(text: str, fallback: str = "") -> str:
    t = clean(text).lower()
    explicit = clean(fallback).upper()
    if explicit and explicit != "AUTO":
        return explicit
    # Amazon BTG sheet names and filenames often contain prefixes like fr-grocery or de-food.
    for code in ["FR", "DE", "IT", "ES", "NL", "PL", "IE", "SE"]:
        c = code.lower()
        if t.startswith(c + "-") or t.startswith(c + "_") or f"/{c}-" in t or f"_{c}_" in t or f"-{c}-" in t:
            return code
        if c in {"fr", "de", "it", "es", "nl", "pl", "ie", "se"} and re.search(rf"(^|[^a-z]){c}([^a-z]|$)", t):
            return code
    return ""


def normalize_btg_path(path: str) -> str:
    text = clean(path)
    # BTG files usually separate path levels with '/', while user's output uses ' > '.
    if "/" in text and ">" not in text:
        parts = [p.strip() for p in text.split("/") if p.strip()]
        return " > ".join(parts)
    return text


def category_name_from_path(path: str) -> str:
    text = clean(path)
    if ">" in text:
        return text.split(">")[-1].strip()
    if "/" in text:
        return text.split("/")[-1].strip()
    return text


def workbook_sheet_names(path: Path) -> List[str]:
    suffix = path.suffix.lower()
    if suffix in {".xlsx", ".xlsm"}:
        wb = load_workbook(path, read_only=True, data_only=True)
        return list(wb.sheetnames)
    if suffix == ".xls":
        if xlrd is None:
            raise ValueError("Za .xls treba paket xlrd. Pokreni: .venv\\Scripts\\python -m pip install xlrd==2.0.1")
        return xlrd.open_workbook(str(path)).sheet_names()
    return [""]


def read_btg_catalog(path: Path, marketplace_hint: str = "") -> List[Dict[str, str]]:
    """
    Čita Amazon Browse Tree Guide datoteke.
    Očekuje sheet s kolonama: Node ID, Node Path, Refinement Link.
    Marketplace se može zadati ručno ili se pokušava zaključiti iz imena sheeta/datoteke
    (npr. fr-grocery, de-food).
    """
    result: List[Dict[str, str]] = []
    suffix = path.suffix.lower()
    if suffix in {".csv", ".txt"}:
        rows = read_csv_rows(path, find_header=True, required=["Node ID", "Node Path"])
        mp = infer_marketplace_from_text(path.name, marketplace_hint)
        for r in rows:
            node_id = clean(r.get("Node ID") or r.get("node_id"))
            raw_path = clean(r.get("Node Path") or r.get("node_path"))
            if node_id and raw_path and mp:
                norm_path = normalize_btg_path(raw_path)
                result.append({"marketplace": mp, "node_id": node_id, "node_path": norm_path, "category_name": category_name_from_path(norm_path), "source": path.name})
        return result

    for sheet in workbook_sheet_names(path):
        low = sheet.lower()
        if low in {"instructions", "refinements", "values", "anleitung"}:
            continue
        try:
            rows = read_generic_rows(path, sheet_name=sheet, find_header=True, required=["Node ID", "Node Path"])
        except Exception:
            continue
        mp = infer_marketplace_from_text(sheet, marketplace_hint) or infer_marketplace_from_text(path.name, marketplace_hint)
        if not mp:
            continue
        for r in rows:
            node_id = clean(r.get("Node ID") or r.get("node_id"))
            raw_path = clean(r.get("Node Path") or r.get("node_path"))
            if not node_id or not raw_path:
                continue
            # Preskoči slučajno krivo pročitane refinement/value tabove
            if node_id.lower() == "node id" or raw_path.lower() == "node path":
                continue
            norm_path = normalize_btg_path(raw_path)
            result.append({
                "marketplace": mp,
                "node_id": node_id,
                "node_path": norm_path,
                "category_name": category_name_from_path(norm_path),
                "source": f"{path.name}::{sheet}",
            })
    return result


def read_category_catalog(path: Path | None, marketplace_hint: str = "") -> List[Dict[str, str]]:
    if path is None or not path.exists() or path.stat().st_size == 0:
        return []

    # 1) Prvo probaj normalizirani catalog format:
    # marketplace,node_id,node_path,category_name
    try:
        rows = read_generic_rows(path, find_header=True, required=["marketplace", "node_id"])
        result: List[Dict[str, str]] = []
        for r in rows:
            mp = clean(r.get("marketplace")).upper() or infer_marketplace_from_text(path.name, marketplace_hint)
            node_id = clean(r.get("node_id") or r.get("Node ID") or r.get("Browse Node ID"))
            category_name = clean(r.get("category_name"))
            node_path = clean(r.get("node_path") or r.get("Node Path"))
            if not category_name and node_path:
                category_name = category_name_from_path(node_path)
            if mp and node_id and (category_name or node_path):
                result.append({"marketplace": mp, "node_id": node_id, "category_name": category_name, "node_path": normalize_btg_path(node_path or category_name), "source": path.name})
        if result:
            return result
    except Exception:
        pass

    # 2) Ako nije normalizirani catalog, tretiraj kao Amazon BTG.
    return read_btg_catalog(path, marketplace_hint=marketplace_hint)


def get_category_index_from_db() -> Dict[Tuple[str, str], Dict[str, str]]:
    idx: Dict[Tuple[str, str], Dict[str, str]] = {}
    if not DB_PATH.exists():
        return idx
    with sqlite3.connect(DB_PATH) as con:
        con.row_factory = sqlite3.Row
        try:
            rows = con.execute("SELECT marketplace, node_id, node_path, category_name, source_file FROM category_catalog").fetchall()
        except Exception:
            return idx
        for item in rows:
            mp = clean(item["marketplace"]).upper()
            node_id = clean(item["node_id"])
            if mp and node_id:
                idx[(mp, node_id)] = {
                    "marketplace": mp,
                    "node_id": node_id,
                    "node_path": clean(item["node_path"]),
                    "category_name": clean(item["category_name"]),
                    "source": clean(item["source_file"]),
                }
    return idx


def import_category_catalog_file(path: Path, marketplace_hint: str = "") -> Dict[str, Any]:
    rows = read_category_catalog(path, marketplace_hint=marketplace_hint)
    now = datetime.utcnow().isoformat(timespec="seconds")
    inserted = 0
    updated = 0
    skipped = 0
    with sqlite3.connect(DB_PATH) as con:
        for r in rows:
            mp = clean(r.get("marketplace")).upper()
            node_id = clean(r.get("node_id"))
            node_path = clean(r.get("node_path"))
            category_name = clean(r.get("category_name")) or category_name_from_path(node_path)
            source_file = clean(r.get("source")) or path.name
            if not mp or not node_id or not node_path:
                skipped += 1
                continue
            existed = con.execute("SELECT 1 FROM category_catalog WHERE marketplace=? AND node_id=?", (mp, node_id)).fetchone() is not None
            con.execute("""
            INSERT INTO category_catalog (marketplace, node_id, node_path, category_name, source_file, imported_at, updated_at)
            VALUES (?, ?, ?, ?, ?, ?, ?)
            ON CONFLICT(marketplace, node_id) DO UPDATE SET
                node_path=excluded.node_path,
                category_name=excluded.category_name,
                source_file=excluded.source_file,
                updated_at=excluded.updated_at
            """, (mp, node_id, node_path, category_name, source_file, now, now))
            if existed:
                updated += 1
            else:
                inserted += 1
        con.commit()
    return {"total_read": len(rows), "inserted": inserted, "updated": updated, "skipped": skipped}


def catalog_stats() -> List[Dict[str, Any]]:
    init_db()
    with sqlite3.connect(DB_PATH) as con:
        con.row_factory = sqlite3.Row
        try:
            items = con.execute("SELECT marketplace, COUNT(*) AS cnt FROM category_catalog GROUP BY marketplace ORDER BY marketplace").fetchall()
        except Exception:
            return []
        return [dict(x) for x in items]



def clear_btg_cache() -> Dict[str, Any]:
    """Briše samo uvezene BTG/category podatke. Learning ručne ispravke ostaju."""
    init_db()
    with sqlite3.connect(DB_PATH) as con:
        try:
            catalog_count = con.execute("SELECT COUNT(*) FROM category_catalog").fetchone()[0]
        except Exception:
            catalog_count = 0
        try:
            file_count = con.execute("SELECT COUNT(*) FROM imported_catalog_files").fetchone()[0]
        except Exception:
            file_count = 0
        con.execute("DELETE FROM category_catalog")
        con.execute("DELETE FROM imported_catalog_files")
        con.commit()
    return {"deleted_category_rows": int(catalog_count), "deleted_imported_files": int(file_count), "stats": catalog_stats()}


def is_supported_catalog_file(path: Path) -> bool:
    return path.is_file() and path.suffix.lower() in {".xls", ".xlsx", ".xlsm", ".csv", ".txt"}


def marketplace_from_folder(path: Path) -> str:
    # Ako korisnik složi folder npr. BTG/FR/fr_garden.xls, parent folder FR je dobar hint.
    for part in reversed(path.parts[:-1]):
        mp = clean(part).upper()
        if mp in {"FR", "DE", "IT", "ES", "NL", "PL", "IE", "SE"}:
            return mp
    return ""


def candidate_project_roots() -> List[Path]:
    roots: List[Path] = []
    for p in [BASE_DIR, *BASE_DIR.parents]:
        if p not in roots:
            roots.append(p)
    return roots


def find_default_btg_folder() -> Path | None:
    # Prvo environment override ako ga korisnik ikad želi. Inače automatski traži folder BTG u projektu.
    env_value = clean(os.environ.get("BTG_FOLDER_PATH", "")) or clean(os.environ.get("BTG_FOLDER", ""))
    if env_value and Path(env_value).exists() and Path(env_value).is_dir():
        return Path(env_value)

    for root in candidate_project_roots():
        for name in ["BTG", "btg", "BrowseTreeGuides", "browse_tree_guides"]:
            candidate = root / name
            if candidate.exists() and candidate.is_dir():
                return candidate
    return None


def find_default_mapping_file() -> Path | None:
    # Prvo environment override, pa poznata imena iz projekta, pa fallback search po nazivu.
    env_value = clean(os.environ.get("MAPPING_FILE_PATH", "")) or clean(os.environ.get("AMAZON_MAPPING_FILE", ""))
    if env_value and Path(env_value).exists() and Path(env_value).is_file():
        return Path(env_value)

    known_names = [
        "mapping_file.xls", "mapping_file.xlsx", "mapping_file.xlsm",
        "european_browse_node_mapping.xls", "european_browse_node_mapping.xlsx",
        "European Browse Node Mapping.xls", "European Browse Node Mapping.xlsx",
    ]
    for root in candidate_project_roots():
        for name in known_names:
            candidate = root / name
            if candidate.exists() and candidate.is_file():
                return candidate

    for root in candidate_project_roots():
        try:
            matches = sorted([
                p for p in root.glob("*mapping*.xls*")
                if p.is_file() and not p.name.startswith("~$")
            ])
            if matches:
                return matches[0]
        except Exception:
            pass
    return None


def project_auto_info() -> Dict[str, str]:
    btg = find_default_btg_folder()
    mapping = find_default_mapping_file()
    return {
        "btg_folder": str(btg) if btg else "",
        "mapping_file": str(mapping) if mapping else "",
    }


def catalog_file_stamp(path: Path) -> Tuple[int, float]:
    st = path.stat()
    return int(st.st_size), float(st.st_mtime)


def catalog_file_is_current(path: Path) -> bool:
    init_db()
    file_path = str(path.resolve())
    try:
        size, modified_time = catalog_file_stamp(path)
        with sqlite3.connect(DB_PATH) as con:
            row = con.execute(
                "SELECT size, modified_time FROM imported_catalog_files WHERE file_path=?",
                (file_path,),
            ).fetchone()
        if not row:
            return False
        return int(row[0]) == size and abs(float(row[1]) - modified_time) < 0.001
    except Exception:
        return False


def mark_catalog_file_imported(path: Path, rows_read: int) -> None:
    init_db()
    size, modified_time = catalog_file_stamp(path)
    now = datetime.utcnow().isoformat(timespec="seconds")
    with sqlite3.connect(DB_PATH) as con:
        con.execute("""
        INSERT INTO imported_catalog_files (file_path, size, modified_time, rows_read, updated_at)
        VALUES (?, ?, ?, ?, ?)
        ON CONFLICT(file_path) DO UPDATE SET
            size=excluded.size,
            modified_time=excluded.modified_time,
            rows_read=excluded.rows_read,
            updated_at=excluded.updated_at
        """, (str(path.resolve()), size, modified_time, int(rows_read), now))
        con.commit()


def marketplace_codes_in_btg_folder(folder_path: Path) -> set[str]:
    codes: set[str] = set()
    if not folder_path.exists() or not folder_path.is_dir():
        return codes
    for file_path in folder_path.rglob("*"):
        if not is_supported_catalog_file(file_path):
            continue
        hint = marketplace_from_folder(file_path) or infer_marketplace_from_text(file_path.name, "")
        if hint:
            codes.add(hint.upper())
    return codes


def catalog_marketplaces_in_db() -> set[str]:
    return {clean(x.get("marketplace")).upper() for x in catalog_stats() if clean(x.get("marketplace"))}


def should_auto_import_btg_folder(folder_path: Path, selected_marketplaces: List[str]) -> bool:
    """
    Koristi BTG cache, ali sigurno: ako se datoteka promijeni po veličini ili modified datumu,
    backend će je ponovno uvesti. Ako su sve potrebne datoteke već uvezene i nisu se mijenjale,
    preskače se sporo čitanje Excela.
    """
    if not folder_path.exists() or not folder_path.is_dir():
        return False

    selected = {m.upper() for m in selected_marketplaces if m}
    selected = {m.upper() for m in (selected_marketplaces or []) if m}
    needed = selected | {"DE"} if selected else set()
    files = []
    for p in folder_path.rglob("*"):
        if not is_supported_catalog_file(p):
            continue
        hint_for_file = marketplace_from_folder(p) or infer_marketplace_from_text(p.name, "")
        hint_for_file = clean(hint_for_file).upper()
        if needed and hint_for_file and hint_for_file not in needed:
            continue
        files.append(p)
    if not files:
        return False

    saw_needed_file = False
    for file_path in files:
        hint = marketplace_from_folder(file_path) or infer_marketplace_from_text(file_path.name, "")
        hint = clean(hint).upper()
        if hint and hint not in (selected | {"DE"}):
            continue
        saw_needed_file = True
        if not catalog_file_is_current(file_path):
            return True

    if not saw_needed_file:
        return False

    folder_mps = marketplace_codes_in_btg_folder(folder_path)
    needed = (selected | {"DE"}) & folder_mps
    existing = catalog_marketplaces_in_db()
    return bool(needed and not needed.issubset(existing))


def import_category_catalog_folder(folder_path: Path, marketplace_hint: str = "AUTO", progress_job_id: str | None = None, progress_start: int = 0, progress_end: int = 30, selected_marketplaces: List[str] | None = None) -> Dict[str, Any]:
    init_db()
    if not folder_path.exists() or not folder_path.is_dir():
        raise ValueError(f"Folder ne postoji ili nije folder: {folder_path}")

    selected = {m.upper() for m in (selected_marketplaces or []) if m}
    needed = selected | {"DE"} if selected else set()
    files = []
    for p in folder_path.rglob("*"):
        if not is_supported_catalog_file(p):
            continue
        hint_for_file = marketplace_from_folder(p) or infer_marketplace_from_text(p.name, "")
        hint_for_file = clean(hint_for_file).upper()
        if needed and hint_for_file and hint_for_file not in needed:
            continue
        files.append(p)
    if not files:
        raise ValueError("U odabranom folderu nisam pronašao .xls/.xlsx/.xlsm/.csv BTG/category datoteke.")

    imported_files: List[Dict[str, Any]] = []
    total_read = inserted = updated = skipped = failed = cached = 0
    sorted_files = sorted(files)
    total_files = max(1, len(sorted_files))

    for index, file_path in enumerate(sorted_files, start=1):
        check_cancelled(progress_job_id)
        percent = progress_start + int(((index - 1) / total_files) * max(1, progress_end - progress_start))
        set_progress(progress_job_id, percent, f"Čitam BTG {index}/{total_files}: {file_path.name}")
        hint = clean(marketplace_hint).upper()
        if not hint or hint == "AUTO":
            hint = marketplace_from_folder(file_path) or "AUTO"
        try:
            if catalog_file_is_current(file_path):
                cached += 1
                imported_files.append({"file": str(file_path), "ok": True, "cached": True, "total_read": 0, "inserted": 0, "updated": 0, "skipped": 0})
                set_progress(progress_job_id, progress_start + int((index / total_files) * max(1, progress_end - progress_start)), f"BTG cache {index}/{total_files}: {file_path.name}")
                continue

            res = import_category_catalog_file(file_path, marketplace_hint=hint)
            total_read += int(res.get("total_read", 0))
            inserted += int(res.get("inserted", 0))
            updated += int(res.get("updated", 0))
            skipped += int(res.get("skipped", 0))
            mark_catalog_file_imported(file_path, int(res.get("total_read", 0)))
            imported_files.append({"file": str(file_path), "ok": True, **res})
            set_progress(progress_job_id, progress_start + int((index / total_files) * max(1, progress_end - progress_start)), f"Uvezen BTG {index}/{total_files}: {file_path.name}")
        except Exception as exc:
            failed += 1
            imported_files.append({"file": str(file_path), "ok": False, "error": str(exc)})
            set_progress(progress_job_id, progress_start + int((index / total_files) * max(1, progress_end - progress_start)), f"Preskočena/greška BTG {index}/{total_files}: {file_path.name}")

    return {
        "folder": str(folder_path),
        "files_found": len(files),
        "files_cached": cached,
        "files_failed": failed,
        "total_read": total_read,
        "inserted": inserted,
        "updated": updated,
        "skipped": skipped,
        "stats": catalog_stats(),
        "files": imported_files[:200],
    }


def build_category_index_from_input(rows: List[Dict[str, str]]) -> Dict[Tuple[str, str], Dict[str, str]]:
    """
    Ako ulazni CSV/XLSX već ima neke provjerene nazive kategorija, iskoristi ih kao lokalni catalog.

    Primjer korisnikovog formata:
      Node Id in FR = 3575160031
      FR = Cuisine et Maison > Produits et accessoires de nettoyage > Seaux

    Ovo NIJE izmišljanje: naziv se koristi samo ako već postoji u ulaznoj datoteci ili kasnije
    u korisnikovoj potvrđenoj/ispravljenoj datoteci.
    """
    idx: Dict[Tuple[str, str], Dict[str, str]] = {}
    for row in rows:
        for mp in ["FR", "IT", "ES"]:
            node_id = clean(row.get(f"Node Id in {mp}")) or clean(row.get(f"{mp}_node_id"))
            value = clean(row.get(mp))
            if not node_id or not value:
                continue
            category_name = value.split(">")[-1].strip() if ">" in value else value
            idx[(mp, node_id)] = {
                "marketplace": mp,
                "node_id": node_id,
                "category_name": category_name,
                "node_path": value,
            }
    return idx


def lookup_category_name(category_index: Dict[Tuple[str, str], Dict[str, str]], marketplace: str, node_id: str) -> Tuple[str, str]:
    rec = category_index.get((marketplace.upper(), clean(node_id)))
    if not rec:
        return "", ""
    path = clean(rec.get("node_path"))
    name = clean(rec.get("category_name"))
    # U glavnom stupcu FR/IT/ES bolje je imati puni path ako postoji.
    return path or name, name


def fill_de_category(row: Dict[str, Any], category_index: Dict[Tuple[str, str], Dict[str, str]]) -> None:
    source_node_id = clean(row.get("NODE ID"))
    row["Node Id in DE"] = source_node_id
    de_path, de_name = lookup_category_name(category_index, "DE", source_node_id) if source_node_id else ("", "")
    row["DE"] = de_path or de_name or clean(row.get("Amazon category name"))


def ensure_output_columns(row: Dict[str, Any], marketplaces: List[str]) -> None:
    row.setdefault("Node Id in DE", "")
    row.setdefault("DE", "")
    for mp in marketplaces:
        row.setdefault(mp, "")
        row.setdefault(f"Node Id in {mp}", "")
        row.setdefault(f"{mp}_node_id", "")
        for suffix in ["status", "note", "confidence", "category_name"]:
            row.setdefault(f"{mp}_{suffix}", "")


def init_db() -> None:
    with sqlite3.connect(DB_PATH) as con:
        con.execute("""
        CREATE TABLE IF NOT EXISTS learning_mappings (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            marketplace TEXT NOT NULL,
            input_signature TEXT NOT NULL,
            source_node_id TEXT,
            source_category_name TEXT,
            product_type TEXT,
            pim_category_name TEXT,
            ean TEXT,
            target_value TEXT NOT NULL,
            target_category_name TEXT,
            target_node_id TEXT,
            status TEXT NOT NULL,
            confidence REAL,
            note TEXT,
            source TEXT,
            created_by TEXT,
            created_at TEXT NOT NULL,
            updated_at TEXT NOT NULL,
            usage_count INTEGER NOT NULL DEFAULT 0,
            UNIQUE(marketplace, input_signature)
        )
        """)
        con.execute("""
        CREATE TABLE IF NOT EXISTS category_catalog (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            marketplace TEXT NOT NULL,
            node_id TEXT NOT NULL,
            node_path TEXT NOT NULL,
            category_name TEXT,
            source_file TEXT,
            imported_at TEXT NOT NULL,
            updated_at TEXT NOT NULL,
            UNIQUE(marketplace, node_id)
        )
        """)
        con.execute("""
        CREATE TABLE IF NOT EXISTS imported_catalog_files (
            file_path TEXT PRIMARY KEY,
            size INTEGER NOT NULL,
            modified_time REAL NOT NULL,
            rows_read INTEGER NOT NULL DEFAULT 0,
            updated_at TEXT NOT NULL
        )
        """)
        # Dodaj target_node_id ako postoji stara baza
        try:
            con.execute("ALTER TABLE learning_mappings ADD COLUMN target_node_id TEXT")
        except Exception:
            pass
        con.commit()


def build_signature(row: Dict[str, Any]) -> str:
    parts = [
        clean(row.get("product type")).lower(),
        clean(row.get("NODE ID")).lower(),
        clean(row.get("Amazon category name")).lower(),
        clean(row.get("category name eng")).lower(),
        clean(row.get("PIM category name")).lower(),
    ]
    return " | ".join(parts)


def find_exact_mapping(row: Dict[str, Any], marketplace: str) -> Dict[str, Any] | None:
    signature = build_signature(row)
    with sqlite3.connect(DB_PATH) as con:
        con.row_factory = sqlite3.Row
        cur = con.execute("SELECT * FROM learning_mappings WHERE marketplace=? AND input_signature=?", (marketplace, signature))
        item = cur.fetchone()
        if item:
            con.execute("UPDATE learning_mappings SET usage_count=usage_count+1 WHERE id=?", (item["id"],))
            con.commit()
            return dict(item)
    return None


def token_sort_ratio(a: str, b: str) -> float:
    at = " ".join(sorted(a.lower().split()))
    bt = " ".join(sorted(b.lower().split()))
    return SequenceMatcher(None, at, bt).ratio() * 100


def find_similar_mapping(row: Dict[str, Any], marketplace: str, min_score: float = 94) -> Tuple[Dict[str, Any] | None, float]:
    signature = build_signature(row)
    best: Dict[str, Any] | None = None
    best_score = 0.0
    with sqlite3.connect(DB_PATH) as con:
        con.row_factory = sqlite3.Row
        for item in con.execute("SELECT * FROM learning_mappings WHERE marketplace=?", (marketplace,)):
            score = token_sort_ratio(signature, item["input_signature"])
            if score > best_score:
                best = dict(item)
                best_score = score
        if best and best_score >= min_score:
            con.execute("UPDATE learning_mappings SET usage_count=usage_count+1 WHERE id=?", (best["id"],))
            con.commit()
            return best, best_score
    return None, best_score


def save_mapping(row: Dict[str, Any], marketplace: str, target_value: str, target_category_name: str, target_node_id: str, confidence: float, status: str, note: str, source: str = "USER_CORRECTION") -> None:
    now = datetime.utcnow().isoformat(timespec="seconds")
    signature = build_signature(row)
    data = {
        "marketplace": marketplace,
        "input_signature": signature,
        "source_node_id": clean(row.get("NODE ID")),
        "source_category_name": clean(row.get("Amazon category name")),
        "product_type": clean(row.get("product type")),
        "pim_category_name": clean(row.get("PIM category name")),
        "ean": clean(row.get("EAN")),
        "target_value": clean(target_value),
        "target_category_name": clean(target_category_name),
        "target_node_id": clean(target_node_id),
        "status": status,
        "confidence": confidence,
        "note": note,
        "source": source,
        "created_by": "winforms_user",
        "now": now,
    }
    with sqlite3.connect(DB_PATH) as con:
        con.execute("""
        INSERT INTO learning_mappings (
            marketplace, input_signature, source_node_id, source_category_name, product_type, pim_category_name, ean,
            target_value, target_category_name, target_node_id, status, confidence, note, source, created_by, created_at, updated_at, usage_count
        ) VALUES (
            :marketplace, :input_signature, :source_node_id, :source_category_name, :product_type, :pim_category_name, :ean,
            :target_value, :target_category_name, :target_node_id, :status, :confidence, :note, :source, :created_by, :now, :now, 0
        )
        ON CONFLICT(marketplace, input_signature) DO UPDATE SET
            target_value=excluded.target_value,
            target_category_name=excluded.target_category_name,
            target_node_id=excluded.target_node_id,
            status=excluded.status,
            confidence=excluded.confidence,
            note=excluded.note,
            source=excluded.source,
            updated_at=excluded.updated_at
        """, data)
        con.commit()


def direct_mapping(mapping_index: Dict[str, Dict[str, str]], category_index: Dict[Tuple[str, str], Dict[str, str]], source_node_id: str, marketplace: str) -> Dict[str, str]:
    node_id = clean(source_node_id)
    if not node_id:
        return {"target_node_id": "", "category_value": "", "status": "MISSING_SOURCE_NODE", "note": "U ulaznom CSV-u fali NODE ID.", "category_name": "", "confidence": "0"}
    record = mapping_index.get(node_id)
    if not record:
        return {"target_node_id": "", "category_value": "", "status": "NO_DIRECT_MAPPING", "note": "NODE ID nije pronađen u MAPPINGS tabu.", "category_name": "", "confidence": "0"}
    target_col = DIRECT_COLUMNS.get(marketplace)
    target_id = clean(record.get(target_col)) if target_col else ""
    if not target_id:
        return {"target_node_id": "", "category_value": "", "status": "MISSING_TARGET_NODE", "note": f"Pronađen source node, ali nema vrijednosti u stupcu '{target_col}'.", "category_name": "", "confidence": "0"}

    category_value, category_name = lookup_category_name(category_index, marketplace, target_id)
    if category_value:
        return {
            "target_node_id": target_id,
            "category_value": category_value,
            "status": "DIRECT_MAPPING_WITH_NAME",
            "note": "Node ID je iz European Browse Node Mapping tablice, a naziv/path iz učitanog category catalog/BTG izvora.",
            "category_name": category_name,
            "confidence": "1",
        }
    return {
        "target_node_id": target_id,
        "category_value": "",
        "status": "DIRECT_NODE_MAPPING_NAME_MISSING",
        "note": f"Node ID je pronađen ({target_id}), ali naziv kategorije nije popunjen jer u lokalnoj BTG/category_catalog bazi nema zapisa za {marketplace} node {target_id}. Uvezi odgovarajući {marketplace} BTG ili provjeri da je BTG folder odabran.",
        "category_name": "",
        "confidence": "0.7",
    }


BTG_FAMILY_ALIASES = {
    "appliances": {"appliances"},
    "automotive": {"automotive", "auto"},
    "baby": {"baby", "baby-products", "babyproducts"},
    "beauty": {"beauty"},
    "books": {"books", "book"},
    "computers": {"computers", "computer"},
    "electronics": {"electronics", "ce", "consumer-electronics"},
    "fashion": {"fashion", "clothing", "shoes"},
    "grocery": {"grocery", "food"},
    "garden": {"garden", "lawn-garden"},
    "health": {"health", "hpc", "drugstore"},
    "industrial": {"industrial"},
    "kitchen": {"kitchen", "home", "home-kitchen"},
    "lighting": {"lighting"},
    "music": {"musical-instruments", "music"},
    "office": {"office-products", "office"},
    "pets": {"pet-supplies", "pets"},
    "software": {"software"},
    "sports": {"sports", "sporting-goods", "sportinggoods"},
    "tools": {"tools", "tools-sgp", "diy"},
    "toys": {"toys", "games"},
    "videogames": {"videogames", "video-games", "video_games"},
}


BTG_CONCEPT_ALIASES = {
    "storage": {"aufbewahren", "aufbewahrung", "ordnen", "ordnung", "storage", "organizing", "organisation", "organization", "organisieren", "opbergen", "przechowywanie", "organizacja", "forvaring", "förvaring", "rangement", "guardar"},
    "waste": {"abfall", "mull", "muell", "müll", "recycling", "trash", "waste", "garbage", "bin", "bins", "dustbin", "smieci", "śmieci", "kosz", "kosze", "odpady", "avfall", "sopor", "soptunna", "soptunnor", "prullenbak", "afval", "dechets", "déchets", "ordures", "basura", "rifiuti"},
    "garden": {"garten", "gartenarbeit", "garden", "gardening", "lawn", "jardin", "jardineria", "jardinería", "giardino", "orto", "ogrod", "ogród", "ogrodowe", "tuin", "trädgård", "tradgard", "trädgårds", "utomhus"},
    "watering": {"bewasserung", "bewässerung", "schlauch", "schlauchsysteme", "watering", "irrigation", "hose", "arrosage", "tuyau", "riego", "manguera", "irrigazione", "tubo", "nawadniania", "nawadnianie", "waz", "wąż", "weze", "węże", "bewatering", "bevattning", "slang"},
    "tank": {"wassertank", "wassertanks", "tank", "tanks", "cistern", "cisterna", "zbiornik", "zbiorniki", "watertank", "watertanks", "vattentank", "vattentankar", "reservoir", "serbatoio", "deposito"},
    "plant_pots": {"blumentopf", "blumentopfe", "pflanzgefass", "pflanzgefäß", "pflanzgefaesse", "pflanzgefäße", "plant", "plants", "planter", "planters", "flowerpot", "flowerpots", "donica", "donice", "kwiat", "kwiaty", "kruka", "krukor", "plantenbak", "bloempot", "pot", "pots", "macetas", "vaso", "vasi"},
    "kitchen": {"kuche", "küche", "kitchen", "kuchnia", "cuisine", "cucina", "cocina", "keuken", "kok", "kök"},
    "cookware": {"kochgeschirr", "kochen", "pfanne", "topf", "cookware", "utensil", "utensils", "naczynia", "garnki", "patelnie", "redskap", "ustensiles", "utensili", "utensilios"},
    "cleaning": {"reinigen", "reinigung", "cleaning", "nettoyage", "limpieza", "pulizia", "sprzatanie", "sprzątanie", "städning", "stadning", "schoonmaak"},
    "tools": {"werkzeug", "werkzeuge", "tools", "narzedzia", "narzędzia", "verktyg", "outils", "herramientas", "attrezzi", "gereedschap"},
    "automotive": {"automotive", "auto", "car", "cars", "fahrzeug", "samochod", "samochód", "bil", "voiture", "coche", "autoaccessoires"},
}


# Spoji ugrađene alias pojmove i rječnik iz category_translation.py.
# category_translation.py je namjerno odvojen da se prijevodi lako mijenjaju bez diranja glavne logike.
for _concept, _aliases in CONCEPT_TRANSLATIONS.items():
    existing = set(BTG_CONCEPT_ALIASES.get(_concept, set()))
    existing.update(_aliases)
    BTG_CONCEPT_ALIASES[_concept] = existing


def concepts_from_text(text: str) -> set[str]:
    norm = " " + normalize_match_text(text) + " "
    concepts: set[str] = set()
    for concept, aliases in BTG_CONCEPT_ALIASES.items():
        for alias in aliases:
            alias_norm = normalize_match_text(alias)
            if not alias_norm:
                continue
            # Za fraze gledamo cijelu frazu, za jednu riječ word-boundary preko razmaka.
            if " " in alias_norm:
                if alias_norm in norm:
                    concepts.add(concept)
                    break
            elif f" {alias_norm} " in norm:
                concepts.add(concept)
                break
    return concepts


def has_blocking_concept_conflict(source_concepts: set[str], target_concepts: set[str]) -> tuple[bool, str]:
    for src in source_concepts:
        blocked = set(OPPOSING_CONCEPTS.get(src, set()))
        hit = blocked & target_concepts
        if hit:
            return True, f"source concept '{src}' conflicts with target concept '{sorted(hit)[0]}'"
    return False, ""


def required_leaf_concepts(source_path_or_text: str) -> set[str]:
    leaf_concepts = concepts_from_text(leaf_name(source_path_or_text))
    critical = set(CRITICAL_LEAF_CONCEPTS)
    return leaf_concepts & critical


def concept_score(source_concepts: set[str], target_concepts: set[str]) -> float:
    if not source_concepts or not target_concepts:
        return 0.0
    overlap = source_concepts & target_concepts
    if not overlap:
        return 0.0
    union = source_concepts | target_concepts
    return 100.0 * len(overlap) / max(1, len(union))


def strip_accents(text: str) -> str:
    return "".join(ch for ch in unicodedata.normalize("NFKD", text) if not unicodedata.combining(ch))


def normalize_match_text(text: str) -> str:
    text = strip_accents(clean(text).lower())
    text = text.replace("&", " and ").replace(">", " ").replace("/", " ").replace("_", "-")
    text = re.sub(r"[^a-z0-9]+", " ", text)
    tokens = [t for t in text.split() if len(t) > 1 and t not in {"and", "the", "for", "with", "und", "der", "die", "das", "les", "des", "los", "las", "gli", "per", "voor", "och", "dla"}]
    return " ".join(tokens)


def canonical_family(value: str) -> str:
    norm = normalize_match_text(value).replace(" ", "-")
    for canonical, aliases in BTG_FAMILY_ALIASES.items():
        for alias in aliases:
            alias_norm = alias.lower().replace("_", "-")
            if alias_norm and (alias_norm == norm or alias_norm in norm):
                return canonical
    return ""


def extract_btg_family(value: str) -> str:
    text = clean(value)
    # Najpouzdanije je ime BTG datoteke, npr. de_garden_browse_tree_guide.xls.
    file_part = Path(text.split("::", 1)[0]).name.lower()
    file_part = re.sub(r"^(de|fr|it|es|nl|pl|ie|se)[_-]", "", file_part)
    file_part = file_part.replace("browse_tree_guide", "").replace("browse-tree-guide", "")
    fam = canonical_family(file_part)
    if fam:
        return fam
    return canonical_family(text)


def path_depth(value: str) -> int:
    text = clean(value)
    if not text:
        return 0
    if ">" in text:
        return len([p for p in text.split(">") if clean(p)])
    if "/" in text:
        return len([p for p in text.split("/") if clean(p)])
    return 1


def leaf_name(value: str) -> str:
    text = clean(value)
    if ">" in text:
        return clean(text.split(">")[-1])
    if "/" in text:
        return clean(text.split("/")[-1])
    return text


def category_records_for_marketplace(category_index: Dict[Tuple[str, str], Dict[str, str]], marketplace: str) -> List[Dict[str, str]]:
    mp = marketplace.upper()
    return [r for (m, _), r in category_index.items() if m == mp]


def btg_similarity_mapping(category_index: Dict[Tuple[str, str], Dict[str, str]], row: Dict[str, Any], marketplace: str) -> Dict[str, str]:
    """
    Sigurniji fallback bez direktnog European mapping stupca.
    Prvo se uzima DE node iz BTG-a, zatim se u ciljnom marketplaceu gleda samo ista BTG obitelj
    (npr. de_garden -> pl_garden). Rezultat se prihvaća samo ako postoji dovoljno siguran
    konceptualni overlap; inače ostaje prazno za ručnu provjeru.
    """
    mp = marketplace.upper().strip()
    source_node_id = clean(row.get("NODE ID"))
    source_rec = category_index.get(("DE", source_node_id)) if source_node_id else None

    source_path = clean(source_rec.get("node_path")) if source_rec else ""
    source_name = clean(source_rec.get("category_name")) if source_rec else ""
    source_file = clean(source_rec.get("source")) if source_rec else ""

    fallback_text = " | ".join(clean(row.get(c)) for c in ["Amazon category name", "category name eng", "PIM category name", "product type"] if clean(row.get(c)))
    query_text = " | ".join(x for x in [source_path, source_name, fallback_text] if x)
    if not query_text:
        return {"target_node_id": "", "category_value": "", "status": "NO_SOURCE_CATEGORY", "note": "Nema DE BTG patha ni naziva kategorije u ulazu za usporedbu.", "category_name": "", "confidence": "0"}

    source_family = extract_btg_family(source_file) or extract_btg_family(clean(row.get("product type"))) or extract_btg_family(query_text)
    all_candidates = category_records_for_marketplace(category_index, mp)
    if not all_candidates:
        return {"target_node_id": "", "category_value": "", "status": "NO_TARGET_BTG", "note": f"Nema učitanih BTG kategorija za {mp}. Ubaci {mp} BTG datoteke u BTG folder.", "category_name": "", "confidence": "0"}

    candidates: List[Dict[str, str]] = []
    wrong_family_count = 0
    for cand in all_candidates:
        cand_path = clean(cand.get("node_path"))
        cand_source = clean(cand.get("source"))
        cand_family = extract_btg_family(cand_source) or extract_btg_family(cand_path)
        if source_family and cand_family and cand_family != source_family:
            wrong_family_count += 1
            continue
        candidates.append(cand)

    if source_family and not candidates:
        return {
            "target_node_id": "",
            "category_value": "",
            "status": "NO_TARGET_BTG_FAMILY",
            "note": f"Za {mp} nema učitane BTG obitelji '{source_family}'. Ne mapiram na druge obitelji da ne dobiješ krivu kategoriju.",
            "category_name": "",
            "confidence": "0",
        }

    source_depth = path_depth(source_path or query_text)
    query_norm = normalize_match_text(query_text)
    leaf_norm = normalize_match_text(leaf_name(source_path or query_text))
    source_concepts = concepts_from_text(" | ".join([source_path, source_name, fallback_text]))
    must_have_concepts = required_leaf_concepts(source_path or source_name or query_text)

    best: Dict[str, str] | None = None
    best_score = -1.0
    best_family = ""
    best_depth = 0
    best_concepts: set[str] = set()
    best_concept_overlap: set[str] = set()

    for cand in candidates:
        cand_path = clean(cand.get("node_path"))
        cand_name = clean(cand.get("category_name"))
        cand_source = clean(cand.get("source"))
        cand_family = extract_btg_family(cand_source) or extract_btg_family(cand_path)
        cand_depth = path_depth(cand_path)
        cand_text = " | ".join([cand_path, cand_name])
        cand_norm = normalize_match_text(cand_text)
        cand_leaf_norm = normalize_match_text(leaf_name(cand_path))
        cand_concepts = concepts_from_text(cand_text)
        overlap = source_concepts & cand_concepts

        # Ako source ima prepoznate pojmove, a kandidat nema nijedan zajednički pojam, ne uzimamo ga.
        # Ovo sprječava slučajeve tipa DE "Mülltonnen" -> PL "Naczynia kuchenne".
        if source_concepts and not overlap:
            continue

        conflict, _conflict_reason = has_blocking_concept_conflict(source_concepts, cand_concepts)
        if conflict:
            continue

        # Ako je zadnji dio DE kategorije jasan i bitan (npr. Mülltonnen, Wassertanks),
        # target mora imati taj isti prevedeni concept. Inače radije ostavljamo prazno.
        if must_have_concepts and not (must_have_concepts & cand_concepts):
            continue

        # Za duboke putanje sa 3+ concepta jedan overlap je često preopćenit.
        # Npr. Garten + Wassertanks ne smije završiti u Donice samo zato što obje imaju "vrt".
        if len(source_concepts) >= 3 and len(overlap) < 2:
            continue

        base_score = token_sort_ratio(query_norm, cand_norm) * 0.16
        leaf_score = token_sort_ratio(leaf_norm, cand_leaf_norm) * 0.10 if leaf_norm and cand_leaf_norm else 0
        family_score = 32 if source_family and cand_family and source_family == cand_family else 0
        depth_score = max(0, 10 - abs(source_depth - cand_depth) * 2.5) if source_depth and cand_depth else 0
        concept_part = concept_score(source_concepts, cand_concepts) * 0.48

        # Mali bonus ako je leaf isti ili vrlo sličan, korisno za IE/engleske BTG-ove.
        leaf_exact_bonus = 8 if leaf_norm and cand_leaf_norm and leaf_norm == cand_leaf_norm else 0
        score = base_score + leaf_score + family_score + depth_score + concept_part + leaf_exact_bonus

        if score > best_score:
            best = cand
            best_score = score
            best_family = cand_family
            best_depth = cand_depth
            best_concepts = cand_concepts
            best_concept_overlap = overlap

    if not best:
        return {
            "target_node_id": "",
            "category_value": "",
            "status": "NO_SAFE_BTG_MATCH",
            "note": f"Nema sigurnog {mp} prijedloga. Kandidati iz krivih BTG obitelji su preskočeni ({wrong_family_count}); traženi leaf concepts={','.join(sorted(must_have_concepts)) or '-'}; source concepts={','.join(sorted(source_concepts)) or '-'}.",
            "category_name": "",
            "confidence": "0",
        }

    # Sa konceptima može proći nešto niži tekstualni score; bez koncepata tražimo puno strožu sličnost.
    threshold = 62 if source_concepts and best_concept_overlap else 88
    target_node_id = clean(best.get("node_id"))
    target_path = clean(best.get("node_path"))
    target_name = clean(best.get("category_name")) or category_name_from_path(target_path)
    confidence = max(0.01, min(0.88, round(best_score / 100, 3)))

    if best_score < threshold:
        return {
            "target_node_id": "",
            "category_value": "",
            "status": "NO_SAFE_BTG_MATCH",
            "note": f"Najbliži {mp} BTG zapis nije dovoljno siguran. Score={round(best_score, 1)}, threshold={threshold}, source_family={source_family or '-'}, best_family={best_family or '-'}, overlap={','.join(sorted(best_concept_overlap)) or '-'}, required={','.join(sorted(must_have_concepts)) or '-'}.",
            "category_name": "",
            "confidence": "0",
        }

    return {
        "target_node_id": target_node_id,
        "category_value": target_path,
        "status": "BTG_SIMILAR_NEEDS_REVIEW",
        "note": f"Prijedlog iz iste BTG obitelji + prijevodni concept match. Score={round(best_score, 1)}, family={best_family or source_family or '-'}, overlap={','.join(sorted(best_concept_overlap)) or '-'}, required={','.join(sorted(must_have_concepts)) or '-'}, depth={source_depth}->{best_depth}. Ručno provjeriti prije potvrde.",
        "category_name": target_name,
        "confidence": str(confidence),
    }


def openai_guess(row: Dict[str, Any], marketplace: str) -> Dict[str, Any]:
    env = load_env()
    api_key = env.get("OPENAI_API_KEY", "")
    model = env.get("OPENAI_MODEL", "gpt-4.1-mini")
    if not api_key:
        return {"target_node_id": "", "category_value": "", "status": "AI_DISABLED", "reason": "OPENAI_API_KEY nije postavljen.", "confidence": 0, "category_name": "", "source_url": ""}

    prompt = f"""
Trebam pomoć za Amazon category/browse node mapping za marketplace {marketplace}.
Vrati strogi JSON s poljima: target_node_id, category_value, category_name, confidence, status, reason, source_url.

PRAVILA:
- Ne izmišljaj ni node ID ni naziv kategorije.
- Ako ne možeš potvrditi da kategorija stvarno postoji na Amazonu ili službenom Amazon/Seller Central izvoru, status mora biti NEED_REVIEW i target_node_id može biti prazan.
- source_url mora biti URL izvora gdje je kategorija pronađena.
- Ako je rezultat samo procjena, status mora biti NEED_REVIEW, ne CONFIRMED.

Ulazni podaci:
product type: {clean(row.get('product type'))}
DE/source NODE ID: {clean(row.get('NODE ID'))}
Amazon category name DE: {clean(row.get('Amazon category name'))}
category name eng: {clean(row.get('category name eng'))}
PIM category name: {clean(row.get('PIM category name'))}
EAN: {clean(row.get('EAN'))}
""".strip()
    payload = {
        "model": model,
        "input": prompt,
        "tools": [{"type": "web_search_preview"}],
        "text": {
            "format": {
                "type": "json_schema",
                "name": "amazon_category_guess",
                "strict": True,
                "schema": {
                    "type": "object",
                    "additionalProperties": False,
                    "properties": {
                        "target_node_id": {"type": "string"},
                        "category_value": {"type": "string"},
                        "category_name": {"type": "string"},
                        "confidence": {"type": "number"},
                        "status": {"type": "string", "enum": ["AI_WEB_MATCH", "MATCH_MEDIUM", "NEED_REVIEW", "NO_MATCH"]},
                        "reason": {"type": "string"},
                        "source_url": {"type": "string"},
                    },
                    "required": ["target_node_id", "category_value", "category_name", "confidence", "status", "reason", "source_url"],
                },
            }
        },
    }
    try:
        req = urllib.request.Request(
            "https://api.openai.com/v1/responses",
            data=json.dumps(payload).encode("utf-8"),
            headers={"Authorization": f"Bearer {api_key}", "Content-Type": "application/json"},
            method="POST",
        )
        with urllib.request.urlopen(req, timeout=60) as resp:
            data = json.loads(resp.read().decode("utf-8"))
        text = ""
        for item in data.get("output", []):
            for content in item.get("content", []):
                if content.get("type") in {"output_text", "text"}:
                    text += content.get("text", "")
        parsed = json.loads(text) if text else {}
        parsed.setdefault("target_node_id", "")
        parsed.setdefault("category_value", "")
        parsed.setdefault("category_name", "")
        parsed.setdefault("confidence", 0)
        parsed.setdefault("status", "NEED_REVIEW")
        parsed.setdefault("reason", "")
        parsed.setdefault("source_url", "")
        # Oprez: AI rezultat uvijek ostavljamo za pregled osim ako korisnik kasnije potvrdi.
        if parsed.get("status") == "AI_WEB_MATCH":
            parsed["status"] = "AI_NEEDS_REVIEW"
            parsed["reason"] = "AI je pronašao mogući izvor, ali rezultat nije službeni import iz BTG/API-ja; ručno potvrditi. " + clean(parsed.get("reason"))
        return parsed
    except Exception as exc:
        return {"target_node_id": "", "category_value": "", "status": "NEED_REVIEW", "reason": f"AI fallback nije uspio: {exc}", "confidence": 0, "category_name": "", "source_url": ""}


def set_marketplace_result(row: Dict[str, Any], mp: str, target_node_id: str, category_value: str, category_name: str, status: str, source: str, note: str, confidence: str | float, source_url: str = "") -> None:
    # Glavni stupac FR/IT/ES = naziv/path kategorije
    row[mp] = category_value
    row[f"{mp}_node_id"] = target_node_id
    row[f"{mp}_category_name"] = category_name or category_value
    row[f"{mp}_status"] = status
    # Source/source_url se ne izbacuju u CSV; ostavljamo samo status/napomenu/confidence.
    row[f"{mp}_note"] = note
    row[f"{mp}_confidence"] = clean(confidence)
    node_col = f"Node Id in {mp}"
    row[node_col] = target_node_id


def process_file(input_path: Path, mapping_path: Path, output_path: Path, marketplaces: List[str], use_ai: bool, overwrite_existing: bool, max_ai_rows: int, category_path: Path | None = None, progress_job_id: str | None = None, progress_start: int = 40, progress_end: int = 98) -> None:
    selected = [m.upper().strip() for m in marketplaces if m.upper().strip() in MARKETPLACES]
    set_progress(progress_job_id, progress_start, "Čitam ulaznu datoteku...")
    rows = read_input_rows(input_path)
    for row in rows:
        drop_unused_marketplace_columns(row)
    set_progress(progress_job_id, progress_start + 4, "Čitam European Browse Node Mapping...")
    mapping_index = build_mapping_index(read_mapping_rows(mapping_path))

    # Nazivi kategorija NE dolaze iz European mapping tablice; ona daje node ID-eve.
    # Zato ih punimo iz lokalne BTG baze, iz inputa i eventualno iz dodatne category datoteke ako je netko pošalje preko API-ja.
    set_progress(progress_job_id, progress_start + 8, "Učitavam BTG/category podatke iz lokalne baze...")
    category_index = get_category_index_from_db()
    category_index.update(build_category_index_from_input(rows))
    if category_path is not None:
        category_index.update(build_category_index(read_category_catalog(category_path)))
    ai_calls = 0
    total_steps = max(1, len(rows) * max(1, len(selected)))
    done_steps = 0

    for row_index, row in enumerate(rows, start=1):
        check_cancelled(progress_job_id)
        ensure_output_columns(row, selected)
        fill_de_category(row, category_index)
        for mp in selected:
            check_cancelled(progress_job_id)
            done_steps += 1
            percent = progress_start + 10 + int((done_steps / total_steps) * max(1, progress_end - progress_start - 12))
            set_progress(progress_job_id, percent, f"Obrađujem red {row_index}/{len(rows)} za {mp}...")
            current_value = clean(row.get(mp))
            if current_value and not overwrite_existing:
                row[f"{mp}_status"] = "ALREADY_FILLED"
                row[f"{mp}_source"] = "INPUT"
                continue

            learned = find_exact_mapping(row, mp)
            if learned:
                set_marketplace_result(
                    row, mp,
                    target_node_id=clean(learned.get("target_node_id")),
                    category_value=clean(learned.get("target_value")),
                    category_name=clean(learned.get("target_category_name")),
                    status="LEARNED_MATCH",
                    source="LEARNING_DB",
                    note=clean(learned.get("note")) or "Pronađeno u learning bazi.",
                    confidence=learned.get("confidence", 1),
                )
                continue

            similar, score = find_similar_mapping(row, mp)
            if similar:
                set_marketplace_result(
                    row, mp,
                    target_node_id=clean(similar.get("target_node_id")),
                    category_value=clean(similar.get("target_value")),
                    category_name=clean(similar.get("target_category_name")),
                    status="LEARNED_SIMILAR",
                    source="LEARNING_DB",
                    note=f"Slično naučeno mapiranje, score={round(score, 1)}. Provjeriti ako je važno.",
                    confidence=round(score / 100, 3),
                )
                continue

            if mp in DIRECT_MARKETPLACES:
                result = direct_mapping(mapping_index, category_index, row.get("NODE ID", ""), mp)
                # Ako mapping za FR/IT/ES postoji, koristi njega jer je sigurniji od sličnosti.
                # Ako ga nema, probaj BTG similarity fallback.
                if result["target_node_id"] or result["status"] in {"MISSING_TARGET_NODE", "DIRECT_NODE_MAPPING_NAME_MISSING"}:
                    set_marketplace_result(
                        row, mp,
                        target_node_id=result["target_node_id"],
                        category_value=result["category_value"],
                        category_name=result["category_name"],
                        status=result["status"],
                        source="EUROPEAN_BROWSE_NODE_MAPPING" if result["status"].startswith("DIRECT") else "MAPPING_TABLE",
                        note=result["note"],
                        confidence=result["confidence"],
                    )
                    continue

            # Za NL/PL/IE/SE i za direktne države bez mappinga: pokušaj iz BTG foldera po sličnosti.
            btg_guess = btg_similarity_mapping(category_index, row, mp)
            if clean(btg_guess.get("target_node_id")):
                set_marketplace_result(
                    row, mp,
                    target_node_id=btg_guess["target_node_id"],
                    category_value=btg_guess["category_value"],
                    category_name=btg_guess["category_name"],
                    status=btg_guess["status"],
                    source="BTG_SIMILARITY",
                    note=btg_guess["note"],
                    confidence=btg_guess["confidence"],
                )
                continue

            if mp in AI_FALLBACK_MARKETPLACES and use_ai and ai_calls < max_ai_rows:
                ai_calls += 1
                guess = openai_guess(row, mp)
                set_marketplace_result(
                    row, mp,
                    target_node_id=clean(guess.get("target_node_id")),
                    category_value=clean(guess.get("category_value")) or clean(guess.get("category_name")),
                    category_name=clean(guess.get("category_name")),
                    status=clean(guess.get("status", "NEED_REVIEW")),
                    source="OPENAI_WEB_SEARCH",
                    note=clean(guess.get("reason")),
                    confidence=clean(guess.get("confidence", 0)),
                    source_url=clean(guess.get("source_url")),
                )
            else:
                set_marketplace_result(
                    row, mp,
                    target_node_id="",
                    category_value="",
                    category_name="",
                    status=btg_guess.get("status", "NEED_REVIEW"),
                    source="BTG_SIMILARITY",
                    note=btg_guess.get("note", "Nema dovoljno podataka za automatski prijedlog."),
                    confidence="0",
                )

    set_progress(progress_job_id, progress_end, "Pišem izlazni CSV...")
    write_csv_rows(rows, output_path, selected)

def save_corrections(rows: List[Dict[str, Any]], marketplaces: List[str]) -> int:
    count = 0
    for row in rows:
        for mp in marketplaces:
            value = clean(row.get(mp))
            node_id = clean(row.get(f"{mp}_node_id")) or clean(row.get(f"Node Id in {mp}"))
            status = clean(row.get(f"{mp}_status"))
            category_name = clean(row.get(f"{mp}_category_name")) or value
            note = clean(row.get(f"{mp}_note"))
            if status not in {"CONFIRMED", "CORRECTED"}:
                continue
            if not value and not category_name:
                continue
            try:
                confidence = float(clean(row.get(f"{mp}_confidence", "1")).replace(",", "."))
            except Exception:
                confidence = 1.0
            save_mapping(row, mp, value or category_name, category_name, node_id, confidence, status, note)
            count += 1
    return count


def _parse_content_disposition(headers: str) -> Tuple[str | None, str | None]:
    cd_line = ""
    for line in headers.splitlines():
        if line.lower().startswith("content-disposition:"):
            cd_line = line
            break
    if not cd_line:
        return None, None
    name = None
    filename = None
    parts = [p.strip() for p in cd_line.split(";")]
    for part in parts:
        if "=" not in part:
            continue
        key, value = part.split("=", 1)
        key = key.strip().lower()
        value = value.strip().strip('"')
        if key == "name":
            name = value
        elif key == "filename":
            filename = value
    return name, filename


def parse_multipart(body: bytes, content_type: str) -> Tuple[Dict[str, str], Dict[str, Tuple[str, bytes]]]:
    m = re.search(r"boundary=(?P<b>[^;]+)", content_type)
    if not m:
        raise ValueError(f"Nedostaje multipart boundary. Content-Type: {content_type}")
    boundary = m.group("b").strip().strip('"').encode("utf-8")
    delimiter = b"--" + boundary
    fields: Dict[str, str] = {}
    files: Dict[str, Tuple[str, bytes]] = {}
    for part in body.split(delimiter):
        part = part.strip(b"\r\n")
        if not part or part == b"--":
            continue
        if b"\r\n\r\n" in part:
            raw_headers, content = part.split(b"\r\n\r\n", 1)
        elif b"\n\n" in part:
            raw_headers, content = part.split(b"\n\n", 1)
        else:
            continue
        headers = raw_headers.decode("utf-8", errors="ignore")
        name, filename = _parse_content_disposition(headers)
        if not name:
            continue
        if content.endswith(b"\r\n"):
            content = content[:-2]
        elif content.endswith(b"\n"):
            content = content[:-1]
        if filename is not None:
            files[name] = (filename, content)
        else:
            fields[name] = content.decode("utf-8", errors="replace")
    return fields, files


class Handler(BaseHTTPRequestHandler):
    server_version = "AmazonCategoryMapperLite/1.3"

    def _json(self, status: int, data: Any) -> None:
        payload = json.dumps(data, ensure_ascii=False).encode("utf-8")
        self.send_response(status)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Content-Length", str(len(payload)))
        self.end_headers()
        self.wfile.write(payload)

    def do_GET(self) -> None:
        parsed = urlparse(self.path)
        if parsed.path == "/health":
            self._json(200, {"status": "ok", "backend": "auto-btg-auto-mapping", **project_auto_info(), "catalog_stats": catalog_stats()})
        elif parsed.path == "/api/progress":
            params = parse_qs(parsed.query)
            job_id = clean((params.get("job_id") or [""])[0])
            self._json(200, get_progress(job_id))
        elif parsed.path == "/api/catalog-stats":
            self._json(200, {"items": catalog_stats()})
        else:
            self._json(404, {"detail": "Not found"})

    def do_POST(self) -> None:
        try:
            length = int(self.headers.get("Content-Length", "0"))
            body = self.rfile.read(length)
            if self.path.startswith("/api/map-csv"):
                self.handle_map_csv(body)
            elif self.path.startswith("/api/import-category-catalog"):
                self.handle_import_category_catalog(body)
            elif self.path.startswith("/api/import-category-folder"):
                self.handle_import_category_folder(body)
            elif self.path.startswith("/api/save-corrections"):
                data = json.loads(body.decode("utf-8"))
                rows = data.get("rows", [])
                marketplaces = [clean(m).upper() for m in data.get("marketplaces", MARKETPLACES)]
                saved = save_corrections(rows, marketplaces)
                self._json(200, {"saved": saved})
            elif self.path.startswith("/api/clear-cache"):
                self._json(200, clear_btg_cache())
            elif self.path.startswith("/api/cancel"):
                try:
                    data = json.loads(body.decode("utf-8")) if body else {}
                except Exception:
                    data = {}
                job_id = clean(data.get("job_id"))
                cancel_job(job_id)
                self._json(200, {"cancelled": True, "job_id": job_id})
            elif self.path.startswith("/api/shutdown"):
                self._json(200, {"status": "shutting_down"})
                threading.Thread(target=self.server.shutdown, daemon=True).start()
            else:
                self._json(404, {"detail": "Not found"})
        except Exception as exc:
            self._json(500, {"detail": str(exc)})

    def handle_import_category_catalog(self, body: bytes) -> None:
        fields, files = parse_multipart(body, self.headers.get("Content-Type", ""))
        category_file = None
        for name in ["category_file", "categoryCatalog", "category_catalog", "btg_file", "catalog_file", "file"]:
            if name in files:
                category_file = files[name]
                break
        if category_file is None:
            available_files = ", ".join(files.keys()) if files else "nema poslanih file polja"
            raise ValueError("Nedostaje BTG/category datoteka. Očekujem file polje 'category_file'. Dobiveno: " + available_files)
        marketplace_hint = clean(fields.get("marketplace", "AUTO")).upper()
        category_name, category_content = category_file
        category_suffix = Path(category_name or "category_catalog.xls").suffix or ".xls"
        category_path = TEMP_DIR / f"import_category_{os.getpid()}_{datetime.utcnow().timestamp()}{category_suffix}"
        category_path.write_bytes(category_content)
        result = import_category_catalog_file(category_path, marketplace_hint=marketplace_hint)
        result["stats"] = catalog_stats()
        self._json(200, result)

    def handle_import_category_folder(self, body: bytes) -> None:
        try:
            data = json.loads(body.decode("utf-8")) if body else {}
        except Exception:
            data = {}
        folder_text = clean(data.get("folder_path")) or clean(data.get("folder"))
        marketplace_hint = clean(data.get("marketplace", "AUTO")).upper()
        if not folder_text:
            raise ValueError("Nedostaje folder_path. Pošalji lokalnu putanju do foldera s BTG datotekama.")
        result = import_category_catalog_folder(Path(folder_text), marketplace_hint=marketplace_hint)
        self._json(200, result)

    def handle_map_csv(self, body: bytes) -> None:
        job_id = ""
        try:
            fields, files = parse_multipart(body, self.headers.get("Content-Type", ""))
            job_id = clean(fields.get("job_id"))
            set_progress(job_id, 1, "Backend je primio datoteku...")

            def pick_file(*names: str):
                for name in names:
                    if name in files:
                        return files[name]
                return None

            input_file = pick_file("input_file", "input", "inputFile", "csv_file", "file")
            mapping_file = pick_file("mapping_file", "mapping", "mappingFile", "mapping_excel", "amazon_mapping")
            category_file = pick_file("category_file", "categoryCatalog", "category_catalog", "btg_file", "catalog_file")
            if input_file is None:
                available_files = ", ".join(files.keys()) if files else "nema poslanih file polja"
                raise ValueError("Nedostaje ulazna datoteka. Pošalji file polje 'input_file'. Može biti CSV/XLSX ili TXT s DE node ID-evima. Dobiveno: " + available_files)

            selected = [m.strip().upper() for m in fields.get("marketplaces", "FR,IT,ES,NL,PL,IE,SE").split(",") if m.strip().upper() in MARKETPLACES]
            use_ai = fields.get("use_ai", "false").lower() == "true" or ai_enabled_from_config()
            overwrite_existing = fields.get("overwrite_existing", "true").lower() == "true"
            btg_folder_text = clean(fields.get("btg_folder_path") or fields.get("btg_folder") or fields.get("category_folder"))
            btg_marketplace_hint = clean(fields.get("btg_marketplace") or fields.get("catalog_marketplace") or "AUTO").upper()
            try:
                max_ai_rows = int(fields.get("max_ai_rows", "50"))
            except Exception:
                max_ai_rows = 50

            check_cancelled(job_id)
            set_progress(job_id, 3, "Spremam privremenu ulaznu datoteku...")
            input_name, input_content = input_file
            input_suffix = Path(input_name or "input.csv").suffix or ".csv"
            input_path = TEMP_DIR / f"input_{os.getpid()}_{datetime.utcnow().timestamp()}{input_suffix}"
            input_path.write_bytes(input_content)

            check_cancelled(job_id)
            set_progress(job_id, 5, "Tražim mapping_file.xls/xlsx...")
            if mapping_file is not None:
                mapping_name, mapping_content = mapping_file
                mapping_suffix = Path(mapping_name or "mapping.xlsx").suffix or ".xlsx"
                mapping_path = TEMP_DIR / f"mapping_{os.getpid()}_{datetime.utcnow().timestamp()}{mapping_suffix}"
                mapping_path.write_bytes(mapping_content)
            else:
                mapping_path = find_default_mapping_file()
                if mapping_path is None:
                    raise ValueError("Ne mogu automatski pronaći Amazon European Browse Node Mapping datoteku. Stavi 'mapping_file.xls' ili '*mapping*.xlsx' u root projekta, pored BTG foldera, ili postavi MAPPING_FILE_PATH environment varijablu.")

            category_path = None
            if category_file is not None:
                category_name, category_content = category_file
                category_suffix = Path(category_name or "category_catalog.csv").suffix or ".csv"
                category_path = TEMP_DIR / f"category_{os.getpid()}_{datetime.utcnow().timestamp()}{category_suffix}"
                category_path.write_bytes(category_content)

            # Automatski odabir cijelog BTG foldera. Korisnik ga više ne mora unositi u WinForms.
            check_cancelled(job_id)
            set_progress(job_id, 7, "Tražim BTG folder...")
            btg_path: Path | None = Path(btg_folder_text) if btg_folder_text else find_default_btg_folder()
            if btg_path is not None and btg_path.exists() and btg_path.is_dir():
                if should_auto_import_btg_folder(btg_path, selected):
                    set_progress(job_id, 10, "Učitavam/obnavljam BTG cache...")
                    import_category_catalog_folder(
                        btg_path,
                        marketplace_hint=btg_marketplace_hint or "AUTO",
                        progress_job_id=job_id,
                        progress_start=10,
                        progress_end=38,
                        selected_marketplaces=selected,
                    )
                else:
                    set_progress(job_id, 38, "BTG cache je već ažuran.")
            else:
                set_progress(job_id, 38, "BTG folder nije pronađen; direktni mapping i dalje radi.")

            check_cancelled(job_id)
            output_path = TEMP_DIR / f"amazon_categories_mapped_{os.getpid()}.csv"
            process_file(
                input_path,
                Path(mapping_path),
                output_path,
                selected,
                use_ai,
                overwrite_existing,
                max_ai_rows,
                category_path=category_path,
                progress_job_id=job_id,
                progress_start=40,
                progress_end=98,
            )
            set_progress(job_id, 100, "Gotovo.", status="complete")
            payload = output_path.read_bytes()
            self.send_response(200)
            self.send_header("Content-Type", "text/csv; charset=utf-8")
            self.send_header("Content-Disposition", 'attachment; filename="amazon_categories_mapped.csv"')
            self.send_header("Content-Length", str(len(payload)))
            self.end_headers()
            self.wfile.write(payload)
        except RuntimeError as exc:
            if "zaustavljena" in str(exc).lower():
                set_progress(job_id, 0, str(exc), status="cancelled")
            else:
                set_progress(job_id, 0, f"Greška: {exc}", status="error")
            raise
        except Exception as exc:
            set_progress(job_id, 0, f"Greška: {exc}", status="error")
            raise

def main() -> None:
    init_db()
    print(f"Backend bez pandas/numpy pokrenut na http://{HOST}:{PORT}")
    print("Verzija: auto BTG folder + auto mapping file + CSV/TXT node ID input")
    print("Zaustavljanje: CTRL+C")
    ThreadingHTTPServer((HOST, PORT), Handler).serve_forever()


if __name__ == "__main__":
    main()
